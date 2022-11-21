import openpyxl
import pandas as pd
from pathlib import Path
from math import isnan
import time
from openpyxl.styles import PatternFill
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

from file_xlsx import FileXlsx

xlsx_file = Path('./Silvery_Port.xlsx')
path_for_new_file = './Povogonka.xlsx'

COLOR_HEADER = 'FFC000'
COLOR_GRAY = 'BDBBB6'
COLOR_PINK = 'E5D1D0'

# Retrieve cell value
listCabin = [57588279, 64130933, 56918501, 56249444, 61893335]


def fill_row(sheet, row, col, color):
    for i in range(1, col):
        sheet.cell(row, i).fill = PatternFill(patternType='solid', fgColor=color)


def create_xlsx_file_header(path, filename, names_list):
    count = len(names_list)
    newxl = FileXlsx(path, filename)
    newxl.row_filling(1, count, names_list)
    newxl.color_row(1, 1, count, COLOR_HEADER)
    newxl.wrap_row(1, 1, count + 1)
    return newxl


def get_values(sheet, rw, colons):
    return [sheet.cell(row=rw, column=colon).value for colon in range(1, colons + 1)]


def fill_rows(sheet, row):
    column = 1
    while sheet.cell(row, column).value:
        sheet.cell(row, column).fill = PatternFill(patternType='solid', fgColor=COLOR_PINK)
        column += 1
    """Add date on the last cell and fill background"""
    sheet.cell(row, column).value = datetime.datetime.now().strftime("%x")
    sheet.cell(row, column).fill = PatternFill(patternType='solid', fgColor=COLOR_GRAY)


def add_new_elements():
    pass



def readXSLX():

    # Load the xlsx file
    # excel_data = pd.read_excel(xlsx_file)
    # # Read the values of the file in the dataframe
    # exel_values = excel_data.to_dict('list')
    # cabin = exel_values.get('Номер вагона')
    # weight = exel_values.get('Фактический вес')
    # if isnan(cabin[-1]):
    #     cabin.pop()
    #     weight.pop()
    #
    # cabin_and_weight = {}
    # count_cabins = len(cabin)
    # if len(cabin) == len(weight):
    #     cabin_and_weight = {cabin[i]: weight[i] for i in range(count_cabins)}

    main_workbook = openpyxl.load_workbook(xlsx_file)  # path to the Excel file
    sheet = main_workbook.active
    rows = sheet.max_row
    columns = sheet.max_column
    header_names = [cell.value for cell in list(sheet.rows)[0] if cell.value]
    header_names.append('Дата послупления')
    print(sheet.print_title_cols)
    """
        Get the values from the main table
        First row and column is header
    """
    values_list = []
    for row in range(2, rows):
        for column in range(1, columns):
            cell = sheet.cell(row, column)
            if cell.value in listCabin:
                values_list += [get_values(sheet, row, column + 1) + [time.strftime("%x")]]

                fill_rows(sheet, row)
    """ Safe changes in the main file """
    main_workbook.save(xlsx_file)

    newxls = create_xlsx_file_header('./', 'test.xlsx', header_names)
    row = 2  # Because we already have a header
    for value in values_list:
        newxls.row_filling(row, len(value), value)
        row += 1



    newxls.save()


if __name__ == "__main__":
    # take_values()
    readXSLX()
