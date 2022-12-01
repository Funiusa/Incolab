from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
from math import isnan
from openpyxl.styles import PatternFill
import time
from collections.abc import Iterable
from file_xlsx import FileXlsx
from os import path
from openpyxl import Workbook

from file_xlsx import FileXlsx

xlsx_file = Path('./Silvery_Port.xlsx')
new_file_path = './test.xlsx'

HEADER = 'FFC000'
GRAY = 'BDBBB6'
PINK = 'E5D1D0'
LIGHT_GREEN = 'B7DEB9'
DEEP_GREEN = '7BC77F'


# def fill_row(sheet, row, col, color):
#     for i in range(1, col):
#         sheet.cell(row, i).fill = PatternFill(patternType='solid', fgColor=color)


# def get_values(wb_sheet, rw, colons):
#     return [wb_sheet.cell(row=rw, column=colon).value for colon in range(1, colons + 1)]


# def fill_rows(wb_sheet, row):
#     """ Filling rows in the main file """
#     column = 1
#     while wb_sheet.cell(row, column).value:
#         wb_sheet.cell(row, column).fill = PatternFill(patternType='solid', fgColor=PINK)
#         column += 1
#     """ Add date on the last cell and fill background """
#     wb_sheet.cell(row, column).value = time.strftime("%x")
#     wb_sheet.cell(row, column).fill = PatternFill(patternType='solid', fgColor=GRAY)


# def get_current_weight_sum(path_to_file, wagons):
#     """ From main file
#         Read the values of the file in the dataframe
#         Convert file in dict for keys and values """
#     excel_data = pd.read_excel(path_to_file)
#     exel_values = excel_data.to_dict('dict')
#     cabin_num = []
#     weight = excel_data.get("Фактический вес")
#     for name, val in exel_values.items():
#         if name == 'Номер вагона':
#             for key, wagon in val.items():
#                 if wagon in wagons:
#                     cabin_num.append(key)
#     return [weight[elem] for elem in cabin_num]


# def add_new_value_in_row(old_row, value):
#     return old_row + (value, )


def create_header(new_xlsx, main_wrk):  # TODO
    """ Create header for new file
        and get the columns number """
    if main_wrk.header_for_newfile is not None:
        new_xlsx.row_filling(1, len(main_wrk.header_for_newfile), main_wrk.header_for_newfile)
        new_xlsx.set_column_header_count(len(main_wrk.header_for_newfile))
        column = new_xlsx.get_column_header_count()
        new_xlsx.row_stretch(column)
        new_xlsx.row_filling(1, column, main_wrk.header_for_newfile)
        new_xlsx.color_row(1, 1, column, HEADER)
        new_xlsx.wrap_row(1, 1, column + 1)
        new_xlsx.border_bottom_row(1, 1, column)
        new_xlsx.save()


def fill_new_xlsx(new_xlsx, values):  # TODO
    """ Getting  Values from main file """
    try:
        row, column = new_xlsx.rows_count(), len(new_xlsx.get_first_row())
        if row == 1:  # Next row after header
            row += 1
        for value in values:
            new_xlsx.clear_color_row(row, column)  # Clear color row
            new_xlsx.row_filling(row, column, value + (time.strftime("%x"), ))  # Adding timestamp in last column
            new_xlsx.rows_count_increment()  # Calculate how much rows we are have
            row += 1
        new_xlsx.save()
    except Exception as e:
        print(f"Error: {e}")


def add_sum_count(new_xlsx, weight_sum):  # TODO
    row, column = new_xlsx.rows_count(), len(new_xlsx.get_first_row())
    if row > 1 and new_xlsx.get_current_rows_count():
        """ Sum """
        new_xlsx.color_row(row, 1, column, LIGHT_GREEN)
        new_xlsx.cell_filling(row, column + 1, "Сумма:")
        new_xlsx.color_cell(row, column + 1, LIGHT_GREEN)
        new_xlsx.cell_filling(row, column + 2, weight_sum)
        """ Count """
        new_xlsx.cell_filling(row, column + 3, "Кол-во:")
        new_xlsx.color_cell(row, column + 3, LIGHT_GREEN)
        new_xlsx.cell_filling(row, column + 4, new_xlsx.get_current_rows_count())
        new_xlsx.border_bottom_row(row, 1, new_xlsx.last_column_count())


def grand_total_sum(file_path, new_xlsx):
    row, column = new_xlsx.rows_count(), len(new_xlsx.get_first_row())
    if row > 1 and new_xlsx.get_current_rows_count():
        """ Sum of all weight """
        exel_data = pd.read_excel(file_path).to_dict('list')
        weight_list = [weight for weight in exel_data.get('Фактический вес')
                       if not isinstance(weight, Iterable) and not isnan(weight)]
        new_xlsx.cell_filling(row + 1, column - 1, "Итого:")
        new_xlsx.cell_filling(row + 1, column, sum(weight_list))
        new_xlsx.color_row(row + 1, 1, column, HEADER, 'mediumGray')


def data_handler(main_file, wagons, newfile_path):
    """ Get the number of columns in first row in main file """
    nbr_columns = len(main_file.get_first_row())
    main_file.set_column_header_count(nbr_columns)
    if path.lexists(newfile_path) is False:
        """ Create the header for the new file if file doesn't exist """
        main_file.header_for_newfile = main_file.get_first_row() + ('Дата послупления', )
    """ Get the data from main file and add in line new element with timestamp. 
        Using wagons list """
    for wagon in wagons:
        row = main_file.get_row(wagon)  # Get row
        if row is False:
            """ If element doesnt exist in main file """
            main_file.nfound_elems.append(wagon)  # If elem not exist add it in list
        elif row and nbr_columns == len(row):
            """ Add new element in list and calculate current sum """
            main_file.received_items_list.append(row)  # tuple
            main_file.current_weight_sum += row[-1]  # Get current weight for new file
        else:
            """ If element already exist with datetime skip and add in list """
            main_file.exists_elems.append(wagon)


def handler_main_file(main_file, wagons):
    """ Color the got row in main file and add the timestamp in the end """
    for wagon in wagons:
        row = main_file.get_row_number(wagon)  # Get the row number
        if row > 0:
            last_column = main_file.get_column_header_count()  # Get the last column
            main_file.color_row(row, 1, last_column, PINK)  # Fill the row
            main_file.cell_filling(row, last_column + 1, time.strftime("%x"))
            main_file.color_cell(row, last_column + 1, GRAY)
    main_file.save()
    main_file.close()


""" TODO check if wagons already was fix the names.
TODO check if list of wagons not empty """
if __name__ == "__main__":

    """ Test values """
    # wagons_list = ()
    # wagons_list = (57588279, 64130933, 56918501, 56249444, 61893335, 3452452)
    wagons_list = (54864947, ) #, 57965113, 61494944, 55760896, 56662430, 60848967, 58061458)

    """ Paths """
    main_file_path = './Silvery_Port.xlsx'
    if not path.lexists(main_file_path):
        print("The main file doesn't exist.")
        exit(-1)

    new_file_path = './test.xlsx'
    """ Get values from main file for new file """
    main_workbook = FileXlsx(main_file_path)  # Create a class
    data_handler(main_workbook, wagons_list, new_file_path)
    handler_main_file(main_workbook, wagons_list)

    """ Adding new values """
    new_workbook = FileXlsx(new_file_path)
    create_header(new_workbook, main_workbook)
    fill_new_xlsx(new_workbook, main_workbook.received_items_list)
    """ Adding sum and count elements """
    add_sum_count(new_workbook, main_workbook.current_weight_sum)
    """ Adding total sum """
    grand_total_sum(new_file_path, new_workbook)
    new_workbook.save()
    new_workbook.close()

    print("Can't find this wagons in main file: ", main_workbook.nfound_elems)
    print("This wagons already exist in new file: ", main_workbook.exists_elems)
    print("Done")

