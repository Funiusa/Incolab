from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


class FileXlsx:
    """ Class for creating new instance of xlsx with parameters """

    def __init__(self, path_to_file, create_empty=False):
        self._path_to_file = path_to_file
        self._workbook = Workbook()
        if create_empty:
            self._workbook.save(path_to_file)

    def save(self):
        """ Save all work """
        self._workbook.save(self._path_to_file)

    def close(self):
        """ Close file """
        self._workbook.close()

    def one_cell_filling(self, row, column, value):
        """ Add element in one cell """
        sheet = self._workbook.active
        try:
            sheet.cell(row, column).value = value
        except ValueError:
            print(f"\nERROR: Cannot convert {value} to Excel. "
                  "Use row/column fillings methods instead.")
        except Exception as e:
            print(f"\nERROR: {e}")

    def wrap_cell(self, row, column):
        """ Wrap cell """
        sheet = self._workbook.active
        try:
            sheet.cell(row, column).alignment = Alignment(wrapText=True, vertical='top', horizontal='center')
        except Exception as e:
            print(f"\nERROR: {e}")

    def wrap_row(self, row, start_column, end_column):
        """ Wrap cells in rows. """
        try:
            for column in range(start_column, end_column):
                self.wrap_cell(row, column)
        except Exception as e:
            print(f"ERROR: {e}")

    def wrap_column(self, column, start_row, end_row):
        """ Wrap cells in columns """
        try:
            for row in range(start_row, end_row):
                self.wrap_cell(row, column)
        except Exception as e:
            print(f"ERROR: {e}")

    def row_filling(self, row, column_end, values):
        """
            Filling cells with values in row.
            All columns and rows must be start at least 1
        """
        try:
            for column in range(column_end):
                self.one_cell_filling(row, column + 1, values[column])
        except Exception as e:
            print(f"ERROR: {e}")

    def column_filling(self, column, start_row, values):
        """
            Filling cells with values in column
            All columns and rows must be start at least 1
        """
        try:
            for value in values:
                self.one_cell_filling(start_row, column, value)
                start_row += 1
        except Exception as e:
            print(f"ERROR: {e}")

    # Create subclass for comments
    def comment_one(self, row, column, comment, authon='Noname'):
        """
            Add comments on the cell
        """
        try:
            sheet = self._workbook.active
            sheet.cell(row, column).comment = Comment(comment, authon)
        except Exception as e:
            print(f"ERROR: {e}")

    def comment_row(self, row, first_column, end_column, comment, author='Noname'):
        """Add comment for cells in row"""
        try:
            for i in range(end_column):
                self.comment_one(row, first_column + i, comment, author)
        except Exception as e:
            print(f"ERROR: {e}")

    def comment_column(self, column, start_row, end_row, comment, author):
        """Add comment for cells in column"""
        try:
            for i in range(end_row - 1):
                self.comment_one(start_row + i, column, comment, author)
        except Exception as e:
            print(f"ERROR: {e}")

    # Create subclass for colors
    def color_one_cell(self, row, column, color, patrn_type='solid'):
        """Add background color for one cell"""
        try:
            sheet = self._workbook.active
            sheet.cell(row, column).fill = PatternFill(patternType=patrn_type, fgColor=color)
        except Exception as e:
            print(f"ERROR: {e}")

    def color_row(self, row, first_column, end_column, color, patrn_type='solid'):
        """Add background color for cells in row"""
        try:
            for i in range(end_column):
                self.color_one_cell(row, first_column + i, color, patrn_type)
        except Exception as e:
            print(f"ERROR: {e}")

    def color_column(self, column, start_row, end_row, color, patrn_type='solid'):
        """Add background color for cells in column"""
        try:
            for i in range(end_row - 1):
                self.color_one_cell(start_row + i, column, color, patrn_type)
        except Exception as e:
            print(f"ERROR: {e}")
